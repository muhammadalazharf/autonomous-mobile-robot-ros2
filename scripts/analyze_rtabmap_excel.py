#!/usr/bin/env python3
"""
analyze_rtabmap_excel.py
========================
Olah hasil ekstraksi RTAB-Map (.xlsx dari export_rtabmap_db.py) jadi
laporan analisis siap-pakai dengan CHART EMBEDDED di Excel.

Input:  ~/hasil_ekstrak/   (output dari export_rtabmap_db.py)
Output: ~/hasil_olahan/    (analisis + grafik)

Untuk setiap database <db>.xlsx, dihitung:
  - kecepatan rata-rata mapping (m/s)
  - density keyframe (node/m)
  - loop closure rate (% dari total node)
  - kecepatan instan per langkah trajectory
  - drift Z (min/max/range)
  - histogram loop closure per node sumber

Lalu disimpan ke <db>_ANALISIS.xlsx dengan sheet:
  RINGKASAN_ANALISIS  - statistik turunan (1 sheet ringkas)
  TRAJECTORY_PLUS     - trajectory + kolom v_instan + jarak kumulatif
  LOOP_CLOSURE_HIST   - histogram count LC per from_id
  GRAFIK              - 4 chart embedded:
                          1. Lintasan 2D (XY scatter)
                          2. Drift Z vs waktu (line)
                          3. Yaw vs waktu (line)
                          4. Kecepatan instan vs waktu (line)

Plus file gabungan PERBANDINGAN_SEMUA.xlsx:
  - Tabel perbandingan semua database dengan metrik turunan
  - Bar chart: loop closure rate per database
  - Bar chart: kecepatan rata-rata per database
  - Scatter: density vs LC rate

CARA PAKAI:
  pip3 install openpyxl
  python3 ~/amr_starter/scripts/analyze_rtabmap_excel.py
  python3 ~/amr_starter/scripts/analyze_rtabmap_excel.py \
      --input ~/hasil_ekstrak --output ~/hasil_olahan
"""
import argparse
import math
import os
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import (BarChart, LineChart, ScatterChart, Reference,
                                Series)
    from openpyxl.chart.trendline import Trendline
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl belum terpasang. Install dulu:")
    print("        pip3 install openpyxl")
    sys.exit(1)


HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')

GOOD_FILL = PatternFill('solid', fgColor='C6EFCE')   # hijau pucat
OK_FILL   = PatternFill('solid', fgColor='FFEB9C')   # kuning pucat
BAD_FILL  = PatternFill('solid', fgColor='FFC7CE')   # merah pucat


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def autosize(ws, maxw=35):
    for col in ws.columns:
        try:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, maxw)
        except Exception:
            pass


def read_sheet(xlsx_path, sheet_name):
    """Baca seluruh sheet sebagai list-of-dicts. Return [] kalau sheet tidak ada."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if r is None or all(v is None for v in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# ----------------------------------------------------------------
# Per-database analysis
# ----------------------------------------------------------------
def analyze_one(xlsx_path, out_dir):
    db_name = Path(xlsx_path).stem
    ring   = read_sheet(xlsx_path, 'RINGKASAN')
    traj   = read_sheet(xlsx_path, 'TRAJECTORY')
    links  = read_sheet(xlsx_path, 'LOOP_CLOSURE')

    # ringkasan original → dict parameter → nilai
    ring_dict = {r.get('parameter'): r.get('nilai') for r in ring}
    total_node = int(safe_float(ring_dict.get('total_node')))
    durasi     = safe_float(ring_dict.get('durasi_recording_s'))
    panjang    = safe_float(ring_dict.get('panjang_lintasan_m'))
    lc_global  = int(safe_float(ring_dict.get('loop_closure_global')))
    lc_lokal   = int(safe_float(ring_dict.get('loop_closure_lokal')))

    if total_node < 2 or len(traj) < 2:
        print(f"    [SKIP] {db_name}: data terlalu sedikit ({total_node} node)")
        return None

    # --- Statistik turunan ---
    v_rata           = (panjang / durasi) if durasi > 0 else 0.0
    density          = (total_node / panjang) if panjang > 0 else 0.0
    lc_rate          = (lc_global / total_node * 100.0) if total_node > 0 else 0.0

    # --- TRAJECTORY_PLUS: tambah kolom v_instan + jarak kumulatif ---
    enriched = []
    cum_dist = 0.0
    prev = None
    z_vals = []
    speeds = []
    for r in traj:
        nid = r.get('node_id')
        t   = safe_float(r.get('t_rel_s'))
        x   = safe_float(r.get('x_m'))
        y   = safe_float(r.get('y_m'))
        z   = safe_float(r.get('z_m'))
        roll  = safe_float(r.get('roll_deg'))
        pitch = safe_float(r.get('pitch_deg'))
        yaw   = safe_float(r.get('yaw_deg'))

        if prev is not None:
            dx = x - prev['x']
            dy = y - prev['y']
            dt = t - prev['t']
            step = math.hypot(dx, dy)
            cum_dist += step
            v_inst = (step / dt) if dt > 0 else 0.0
        else:
            v_inst = 0.0

        z_vals.append(z)
        speeds.append(v_inst)
        enriched.append([nid, t, x, y, z, roll, pitch, yaw,
                         round(v_inst, 4), round(cum_dist, 4)])
        prev = {'x': x, 'y': y, 't': t}

    # Z drift stats
    z_min  = min(z_vals) if z_vals else 0.0
    z_max  = max(z_vals) if z_vals else 0.0
    z_range = z_max - z_min
    v_max   = max(speeds) if speeds else 0.0
    v_avg_inst = (sum(speeds) / len(speeds)) if speeds else 0.0

    # --- LC histogram ---
    lc_global_links = [l for l in links if l.get('type_id') == 2]
    counter = Counter(l.get('from_id') for l in lc_global_links)
    hist = sorted(counter.items(), key=lambda kv: -kv[1])

    # --- Klasifikasi kualitas ---
    if lc_rate >= 30:
        quality = 'SANGAT BAIK'
        q_fill = GOOD_FILL
    elif lc_rate >= 10:
        quality = 'BAIK'
        q_fill = OK_FILL
    elif lc_rate >= 1:
        quality = 'CUKUP'
        q_fill = OK_FILL
    else:
        quality = 'KURANG (linear, sedikit revisit)'
        q_fill = BAD_FILL

    # --- Build workbook ---
    wb = Workbook()

    # Sheet 1: RINGKASAN_ANALISIS
    ws = wb.active
    ws.title = 'RINGKASAN_ANALISIS'
    ws.append(['metrik', 'nilai', 'satuan', 'interpretasi'])
    style_header(ws)
    rows = [
        ['database',               db_name,                    '',     ''],
        ['total_node (keyframe)',  total_node,                 'node', ''],
        ['durasi_recording',       round(durasi, 2),           's',    ''],
        ['panjang_lintasan',       round(panjang, 3),          'm',    ''],
        ['', '', '', ''],
        ['KECEPATAN RATA-RATA',    round(v_rata, 4),           'm/s',
         f'≈ {v_rata*3.6:.2f} km/h'],
        ['kecepatan_instan_rata',  round(v_avg_inst, 4),       'm/s',
         'rata-rata dari Δjarak / Δt'],
        ['kecepatan_instan_max',   round(v_max, 4),            'm/s',  ''],
        ['', '', '', ''],
        ['DENSITY KEYFRAME',       round(density, 3),          'node/m',
         f'{density:.1f} keyframe per meter'],
        ['', '', '', ''],
        ['loop_closure_global',    lc_global,                  'count', ''],
        ['loop_closure_lokal',     lc_lokal,                   'count', ''],
        ['LOOP CLOSURE RATE',      round(lc_rate, 2),          '%',
         '<5% kurang | 10-30% baik | >30% sangat baik'],
        ['kualitas_peta',          quality,                    '',     ''],
        ['', '', '', ''],
        ['z_min',                  round(z_min, 4),            'm',    ''],
        ['z_max',                  round(z_max, 4),            'm',    ''],
        ['DRIFT Z (z_max − z_min)', round(z_range, 4),         'm',
         'lantai datar harusnya <0.1 m'],
    ]
    for r in rows:
        ws.append(r)
    # warna baris kualitas
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label and 'LOOP CLOSURE RATE' in str(label):
            ws.cell(row=row_idx, column=2).fill = q_fill
        if label == 'kualitas_peta':
            ws.cell(row=row_idx, column=2).fill = q_fill
        if label and 'DRIFT Z' in str(label):
            ws.cell(row=row_idx, column=2).fill = (GOOD_FILL if z_range < 0.1
                                                  else OK_FILL if z_range < 0.5
                                                  else BAD_FILL)
    autosize(ws)

    # Sheet 2: TRAJECTORY_PLUS
    ws2 = wb.create_sheet('TRAJECTORY_PLUS')
    ws2.append(['node_id', 't_rel_s', 'x_m', 'y_m', 'z_m',
                'roll_deg', 'pitch_deg', 'yaw_deg',
                'v_instan_mps', 'jarak_kumulatif_m'])
    style_header(ws2)
    for r in enriched:
        ws2.append(r)
    autosize(ws2, maxw=15)

    # Sheet 3: LOOP_CLOSURE_HIST
    ws3 = wb.create_sheet('LOOP_CLOSURE_HIST')
    ws3.append(['from_id', 'jumlah_LC_keluar'])
    style_header(ws3)
    for nid, cnt in hist:
        ws3.append([nid, cnt])
    autosize(ws3)

    # Sheet 4: GRAFIK
    wsg = wb.create_sheet('GRAFIK')
    wsg.cell(row=1, column=1, value=f'Grafik analisis — {db_name}').font = Font(bold=True, size=14)

    n_data = len(enriched)
    end_row = n_data + 1  # baris terakhir di TRAJECTORY_PLUS (header di row 1)

    # 1. Lintasan 2D (XY Scatter)
    ch1 = ScatterChart()
    ch1.title = '1. Lintasan 2D (x vs y)'
    ch1.x_axis.title = 'x (m)'
    ch1.y_axis.title = 'y (m)'
    xs = Reference(ws2, min_col=3, min_row=2, max_row=end_row)
    ys = Reference(ws2, min_col=4, min_row=2, max_row=end_row)
    series1 = Series(ys, xs, title='lintasan')
    ch1.series.append(series1)
    ch1.height = 12
    ch1.width  = 18
    wsg.add_chart(ch1, 'A3')

    # 2. Drift Z vs waktu (line)
    ch2 = LineChart()
    ch2.title = '2. Drift Z vs Waktu'
    ch2.x_axis.title = 't (s)'
    ch2.y_axis.title = 'z (m)'
    z_data = Reference(ws2, min_col=5, min_row=1, max_row=end_row)
    t_data = Reference(ws2, min_col=2, min_row=2, max_row=end_row)
    ch2.add_data(z_data, titles_from_data=True)
    ch2.set_categories(t_data)
    ch2.height = 10
    ch2.width  = 18
    wsg.add_chart(ch2, 'A28')

    # 3. Yaw vs waktu (line)
    ch3 = LineChart()
    ch3.title = '3. Yaw vs Waktu'
    ch3.x_axis.title = 't (s)'
    ch3.y_axis.title = 'yaw (deg)'
    yaw_data = Reference(ws2, min_col=8, min_row=1, max_row=end_row)
    ch3.add_data(yaw_data, titles_from_data=True)
    ch3.set_categories(t_data)
    ch3.height = 10
    ch3.width  = 18
    wsg.add_chart(ch3, 'A50')

    # 4. Kecepatan instan vs waktu (line)
    ch4 = LineChart()
    ch4.title = '4. Kecepatan Instan vs Waktu'
    ch4.x_axis.title = 't (s)'
    ch4.y_axis.title = 'v (m/s)'
    v_data = Reference(ws2, min_col=9, min_row=1, max_row=end_row)
    ch4.add_data(v_data, titles_from_data=True)
    ch4.set_categories(t_data)
    ch4.height = 10
    ch4.width  = 18
    wsg.add_chart(ch4, 'A72')

    # 5. Histogram LC (bar) — kalau ada data
    if hist:
        ch5 = BarChart()
        ch5.type = 'col'
        ch5.title = '5. Histogram Loop Closure per from_id (top 50)'
        ch5.x_axis.title = 'from_id'
        ch5.y_axis.title = 'jumlah LC'
        max_h = min(len(hist) + 1, 51)
        hcat = Reference(ws3, min_col=1, min_row=2, max_row=max_h)
        hval = Reference(ws3, min_col=2, min_row=1, max_row=max_h)
        ch5.add_data(hval, titles_from_data=True)
        ch5.set_categories(hcat)
        ch5.height = 10
        ch5.width  = 18
        wsg.add_chart(ch5, 'A94')

    out_path = Path(out_dir) / f'{db_name}_ANALISIS.xlsx'
    wb.save(out_path)

    return {
        'database':           db_name,
        'total_node':         total_node,
        'durasi_s':           round(durasi, 2),
        'panjang_m':          round(panjang, 3),
        'v_rata_mps':         round(v_rata, 4),
        'v_max_inst':         round(v_max, 4),
        'density_node_per_m': round(density, 3),
        'loop_closure':       lc_global,
        'lc_rate_persen':     round(lc_rate, 2),
        'kualitas':           quality,
        'z_drift_m':          round(z_range, 4),
        'file':               out_path.name,
    }


# ----------------------------------------------------------------
# Workbook gabungan: perbandingan semua DB
# ----------------------------------------------------------------
def build_summary(results, out_dir):
    headers = ['database', 'total_node', 'durasi_s', 'panjang_m',
               'v_rata_mps', 'v_max_inst', 'density_node_per_m',
               'loop_closure', 'lc_rate_persen', 'kualitas',
               'z_drift_m', 'file']

    # urutkan: lc_rate descending
    results = sorted(results, key=lambda r: -r['lc_rate_persen'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'PERBANDINGAN'
    ws.append(headers)
    style_header(ws)

    for r in results:
        row = [r[h] for h in headers]
        ws.append(row)

    # Warnai kolom kualitas + lc_rate
    for ridx in range(2, ws.max_row + 1):
        q_cell = ws.cell(row=ridx, column=10)  # kualitas col 10
        if q_cell.value == 'SANGAT BAIK':
            q_cell.fill = GOOD_FILL
        elif q_cell.value == 'BAIK':
            q_cell.fill = GOOD_FILL
        elif q_cell.value == 'CUKUP':
            q_cell.fill = OK_FILL
        else:
            q_cell.fill = BAD_FILL

    autosize(ws, maxw=22)

    # Sheet GRAFIK gabungan
    wsg = wb.create_sheet('GRAFIK')
    wsg.cell(row=1, column=1, value='Perbandingan kualitas mapping antar database').font = Font(bold=True, size=14)

    end_row = len(results) + 1
    db_ref = Reference(ws, min_col=1, min_row=2, max_row=end_row)

    # Bar chart 1: LC rate per database
    bc1 = BarChart()
    bc1.type = 'bar'
    bc1.title = 'Loop Closure Rate (%) per database'
    bc1.x_axis.title = 'LC rate (%)'
    bc1.y_axis.title = 'database'
    lcref = Reference(ws, min_col=9, min_row=1, max_row=end_row)
    bc1.add_data(lcref, titles_from_data=True)
    bc1.set_categories(db_ref)
    bc1.height = 14
    bc1.width  = 22
    wsg.add_chart(bc1, 'A3')

    # Bar chart 2: kecepatan rata-rata
    bc2 = BarChart()
    bc2.type = 'bar'
    bc2.title = 'Kecepatan rata-rata mapping (m/s) per database'
    bc2.x_axis.title = 'v (m/s)'
    vref = Reference(ws, min_col=5, min_row=1, max_row=end_row)
    bc2.add_data(vref, titles_from_data=True)
    bc2.set_categories(db_ref)
    bc2.height = 14
    bc2.width  = 22
    wsg.add_chart(bc2, 'A33')

    # Bar chart 3: panjang lintasan
    bc3 = BarChart()
    bc3.type = 'bar'
    bc3.title = 'Panjang lintasan (m) per database'
    bc3.x_axis.title = 'lintasan (m)'
    pref = Reference(ws, min_col=4, min_row=1, max_row=end_row)
    bc3.add_data(pref, titles_from_data=True)
    bc3.set_categories(db_ref)
    bc3.height = 14
    bc3.width  = 22
    wsg.add_chart(bc3, 'A63')

    # Bar chart 4: drift Z
    bc4 = BarChart()
    bc4.type = 'bar'
    bc4.title = 'Drift Z (m) per database — lebih kecil lebih baik'
    bc4.x_axis.title = 'drift z (m)'
    zref = Reference(ws, min_col=11, min_row=1, max_row=end_row)
    bc4.add_data(zref, titles_from_data=True)
    bc4.set_categories(db_ref)
    bc4.height = 14
    bc4.width  = 22
    wsg.add_chart(bc4, 'A93')

    out_path = Path(out_dir) / 'PERBANDINGAN_SEMUA_ANALISIS.xlsx'
    wb.save(out_path)
    return out_path


# ----------------------------------------------------------------
# main
# ----------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--input',  default=os.path.expanduser('~/hasil_ekstrak'),
                    help='Folder hasil ekstraksi .xlsx (default: ~/hasil_ekstrak)')
    ap.add_argument('--output', default=os.path.expanduser('~/hasil_olahan'),
                    help='Folder output analisis (default: ~/hasil_olahan)')
    args = ap.parse_args()

    in_dir  = os.path.expanduser(args.input)
    out_dir = os.path.expanduser(args.output)
    os.makedirs(out_dir, exist_ok=True)

    xlsx_files = sorted(Path(in_dir).glob('*.xlsx'))
    # skip file ringkasan gabungan
    xlsx_files = [p for p in xlsx_files
                  if 'RINGKASAN_SEMUA' not in p.name
                  and 'PERBANDINGAN' not in p.name]
    if not xlsx_files:
        print(f"[ERROR] tidak ada .xlsx di {in_dir}")
        sys.exit(1)

    print('=' * 70)
    print('  ANALISIS RTAB-MAP — turunan statistik + chart embedded')
    print(f'  Input : {in_dir}   ({len(xlsx_files)} database)')
    print(f'  Output: {out_dir}')
    print('=' * 70)

    results = []
    for xp in xlsx_files:
        print(f'\n[→] {xp.name}')
        res = analyze_one(xp, out_dir)
        if res:
            results.append(res)
            print(f'    [✓] v={res["v_rata_mps"]} m/s | '
                  f'density={res["density_node_per_m"]} node/m | '
                  f'LC rate={res["lc_rate_persen"]}% '
                  f'({res["kualitas"]})')

    if not results:
        print('\n[ERROR] Tidak ada database yang berhasil dianalisis.')
        sys.exit(1)

    summary_path = build_summary(results, out_dir)

    print('\n' + '=' * 70)
    print(f'  SELESAI — {len(results)} database dianalisis')
    print(f'  Gabungan: {summary_path}')
    print('\n  Top-5 database dengan LC rate tertinggi:')
    top5 = sorted(results, key=lambda r: -r['lc_rate_persen'])[:5]
    print(f'  {"Database":<32} {"LC%":>7} {"v(m/s)":>8} {"Kualitas":<20}')
    print(f'  {"-"*32} {"-"*7} {"-"*8} {"-"*20}')
    for r in top5:
        print(f'  {r["database"][:32]:<32} '
              f'{r["lc_rate_persen"]:>6}% '
              f'{r["v_rata_mps"]:>8} '
              f'{r["kualitas"]:<20}')
    print('=' * 70)
    print(f'\n  Buka di Excel: {out_dir}/<db>_ANALISIS.xlsx → sheet GRAFIK')


if __name__ == '__main__':
    main()
