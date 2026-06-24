#!/usr/bin/env python3
"""
make_master_recap.py  (versi proyek AMR — auto-isi Excel)
=========================================================
Gabungkan seluruh hasil uji (odometry, localization, navigation, topic
snapshot) menjadi:
  1. amr_test_recap.csv         — semua baris mentah jadi satu (universal)
  2. amr_test_recap.xlsx        — multi-sheet SIAP-LAPORAN:
       - Odometry_Raw      + Odometry_Summary    (tabel + agregat error)
       - Localization_Raw  + Localization_Summary
       - Navigation_Raw    + Navigation_Summary
       - Topic_Snapshot
       - Evidence_Index    (checklist bukti)

File .xlsx inilah yang tinggal dibuka di Excel dan disalin ke laporan.

Cara pakai:
  python3 scripts/make_master_recap.py --input ~/amr_test_results

Dependensi .xlsx: openpyxl (pip3 install openpyxl).
Tanpa openpyxl tetap menghasilkan CSV.
"""
import argparse
import csv
import os

from ros_helpers import ensure_dir, now_iso, safe_float


# ---------------------------------------------------------------- helper
def _is_num(v):
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


# ---------------------------------------------------------------- baca CSV
def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------- CSV gabungan
def write_master_csv(out_path, odom, loc, nav, snap):
    all_rows = []
    for r in odom:
        all_rows.append({'test_type': 'odometry_validation', **r})
    for r in loc:
        all_rows.append({'test_type': 'localization_to_map', **r})
    for r in nav:
        all_rows.append({'test_type': 'navigation_demo', **r})
    for r in snap:
        all_rows.append({'test_type': 'topic_snapshot', **r})
    if not all_rows:
        return 0

    preferred = ['test_type', 'timestamp', 'trial', 'status', 'notes']
    keys = [k for k in preferred if any(k in r for r in all_rows)]
    for r in all_rows:
        for k in r:
            if k not in keys:
                keys.append(k)

    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in all_rows:
            w.writerow(r)
    return len(all_rows)


# ---------------------------------------------------------------- Excel
def build_excel(out_path, odom, loc, nav, snap, raw_dir=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    HF = PatternFill('solid', fgColor='1F4E79')
    HFONT = Font(bold=True, color='FFFFFF')
    TITLE = Font(bold=True, size=13)
    SUBFILL = PatternFill('solid', fgColor='DDEBF7')
    GOOD = PatternFill('solid', fgColor='C6EFCE')
    OK = PatternFill('solid', fgColor='FFEB9C')
    BAD = PatternFill('solid', fgColor='FFC7CE')
    CENTER = Alignment(horizontal='center')

    wb = Workbook()
    wb.remove(wb.active)

    def style_header(ws, row=1):
        for c in ws[row]:
            c.font = HFONT
            c.fill = HF
            c.alignment = CENTER

    def autosize(ws, maxw=28):
        for col in ws.columns:
            cells = [c for c in col if c.value is not None]
            if not cells:
                continue
            w = max(len(str(c.value)) for c in cells)
            ws.column_dimensions[get_column_letter(cells[0].column)].width = min(w + 2, maxw)

    def dump_raw(name, rows):
        ws = wb.create_sheet(name)
        if not rows:
            ws['A1'] = '(belum ada data — jalankan logger dulu)'
            return ws
        headers = list(rows[0].keys())
        ws.append(headers)
        style_header(ws)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])
        autosize(ws)
        return ws

    # chart helper (openpyxl native — tidak butuh matplotlib)
    from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series

    # ---- ODOMETRY ----
    dump_raw('Odometry_Raw', odom)
    ws = wb.create_sheet('Odometry_Summary')
    ws['A1'] = 'Tabel Laporan: Validasi Odometry'
    ws['A1'].font = TITLE
    ws.append([])
    head = ['Trial', 'Jarak Aktual (m)', 'Jarak Odometry (m)',
            'Selisih (m)', 'Error (%)']
    ws.append(head)
    odom_hdr_row = ws.max_row
    style_header(ws, ws.max_row)
    errs = []
    for r in odom:
        ep = safe_float(r.get('error_pct'))
        errs.append(ep)
        ws.append([r.get('trial', ''),
                   safe_float(r.get('actual_distance_m')),
                   safe_float(r.get('odom_straight_distance_m')),
                   safe_float(r.get('error_m')),
                   ep])
        cell = ws.cell(row=ws.max_row, column=5)
        cell.fill = GOOD if ep < 5 else OK if ep < 10 else BAD
    if errs:
        ws.append([])
        mean_err = sum(errs) / len(errs)
        ws.append(['RATA-RATA ERROR (%)', '', '', '', round(mean_err, 3)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
        ws.append(['ERROR MIN / MAX (%)', '', '', round(min(errs), 3), round(max(errs), 3)])
        ws.append(['JUMLAH TRIAL', '', '', '', len(errs)])
        ws.append([])
        ws.append(['Interpretasi: error <5% sangat baik, 5-10% baik, >10% perlu kalibrasi ulang PPR'])
    autosize(ws)

    # chart perbandingan odometry
    if odom:
        data_end = odom_hdr_row + len(odom)
        cats = Reference(ws, min_col=1, min_row=odom_hdr_row + 1, max_row=data_end)
        # Bar: aktual vs odom
        bc = BarChart()
        bc.type = 'col'
        bc.title = 'Perbandingan Jarak Aktual vs Odometry'
        bc.y_axis.title = 'jarak (m)'
        bc.x_axis.title = 'trial'
        d = Reference(ws, min_col=2, max_col=3, min_row=odom_hdr_row, max_row=data_end)
        bc.add_data(d, titles_from_data=True)
        bc.set_categories(cats)
        bc.height = 8; bc.width = 16
        ws.add_chart(bc, 'H3')
        # Line: error %
        lc = LineChart()
        lc.title = 'Error (%) per Trial'
        lc.y_axis.title = 'error (%)'
        lc.x_axis.title = 'trial'
        de = Reference(ws, min_col=5, min_row=odom_hdr_row, max_row=data_end)
        lc.add_data(de, titles_from_data=True)
        lc.set_categories(cats)
        lc.height = 8; lc.width = 16
        ws.add_chart(lc, 'H20')

    # ---- LOCALIZATION ----
    dump_raw('Localization_Raw', loc)
    ws = wb.create_sheet('Localization_Summary')
    ws['A1'] = 'Tabel Laporan: Localization terhadap Peta'
    ws['A1'].font = TITLE
    ws.append([])
    head = ['Trial', 'Peta (.db)', 'Jumlah Pose', 'Rate (Hz)',
            'Status', 'x (m)', 'y (m)', 'yaw (rad)']
    ws.append(head)
    style_header(ws, ws.max_row)
    for r in loc:
        status = r.get('status', '')
        ws.append([r.get('trial', ''), r.get('map_file', ''),
                   r.get('pose_message_count', ''),
                   safe_float(r.get('estimated_rate_hz')),
                   status,
                   r.get('last_x_m', ''), r.get('last_y_m', ''),
                   r.get('last_yaw_rad', '')])
        cell = ws.cell(row=ws.max_row, column=5)
        cell.fill = GOOD if 'LOCK' in str(status) or 'RECEIVED' in str(status) else BAD
    ws.append([])
    ws.append(['Interpretasi: status LOCK/POSE_RECEIVED = robot berhasil mengenali posisi di peta'])
    autosize(ws)

    # ---- NAVIGATION ----
    dump_raw('Navigation_Raw', nav)
    ws = wb.create_sheet('Navigation_Summary')
    ws['A1'] = 'Tabel Laporan: Mode Demo Navigation2'
    ws['A1'].font = TITLE
    ws.append([])
    head = ['Trial', 'Start→Goal', 'Global Path', 'cmd_vel Aktif',
            'Robot Bergerak', 'Displacement (m)', 'v_max (m/s)', 'Status']
    ws.append(head)
    style_header(ws, ws.max_row)
    for r in nav:
        moved = r.get('robot_moved', '')
        ws.append([r.get('trial', ''),
                   f"{r.get('start_label', '')}→{r.get('goal_label', '')}",
                   r.get('global_path_received', ''),
                   r.get('cmd_vel_active', ''),
                   moved,
                   safe_float(r.get('odom_displacement_m')),
                   safe_float(r.get('max_linear_x_abs_mps')),
                   r.get('status', '')])
        cell = ws.cell(row=ws.max_row, column=5)
        cell.fill = GOOD if str(moved).upper() == 'YA' else BAD
    ws.append([])
    ws.append(['Interpretasi: Global Path=YA + cmd_vel Aktif=YA + Robot Bergerak=YA → demo navigasi berhasil'])
    autosize(ws)

    # ---- RAW TIME-SERIES (per trial) + chart trajektori ----
    if raw_dir and os.path.isdir(raw_dir):
        raw_files = sorted(f for f in os.listdir(raw_dir) if f.endswith('.csv'))
        for fname in raw_files:
            rows = read_csv(os.path.join(raw_dir, fname))
            if not rows:
                continue
            # nama sheet: RAW_<file tanpa ekstensi>, batas 31 char Excel
            sname = ('RAW_' + fname[:-4])[:31]
            ws = wb.create_sheet(sname)
            headers = list(rows[0].keys())
            ws.append(headers)
            style_header(ws)
            for r in rows:
                ws.append([safe_float(r[h]) if _is_num(r[h]) else r[h]
                           for h in headers])
            autosize(ws, maxw=16)
            # chart trajektori x-y untuk file pose/odom
            if 'x_m' in headers and 'y_m' in headers:
                end = len(rows) + 1
                ix = headers.index('x_m') + 1
                iy = headers.index('y_m') + 1
                sc = ScatterChart()
                sc.title = f'Lintasan {fname[:-4]} (x vs y)'
                sc.x_axis.title = 'x (m)'; sc.y_axis.title = 'y (m)'
                xs = Reference(ws, min_col=ix, min_row=2, max_row=end)
                ys = Reference(ws, min_col=iy, min_row=2, max_row=end)
                sc.series.append(Series(ys, xs, title='lintasan'))
                sc.height = 10; sc.width = 14
                anchor = get_column_letter(len(headers) + 2) + '2'
                ws.add_chart(sc, anchor)

    # ---- TOPIC SNAPSHOT ----
    dump_raw('Topic_Snapshot', snap)

    # ---- EVIDENCE INDEX ----
    ws = wb.create_sheet('Evidence_Index')
    ws.append(['No', 'Nama Bukti', 'Jenis Bukti', 'Subbab Laporan', 'Status', 'Catatan'])
    style_header(ws)
    checklist = [
        [1, 'odometry_trials.csv', 'Data CSV', '2.9 Kalibrasi Odometry',
         'OTOMATIS' if odom else 'BELUM', f'{len(odom)} trial'],
        [2, 'localization_trials.csv', 'Data CSV', '2.13 Localization',
         'OTOMATIS' if loc else 'BELUM', f'{len(loc)} trial'],
        [3, 'navigation_trials.csv', 'Data CSV', '2.14 Navigation2',
         'OTOMATIS' if nav else 'BELUM', f'{len(nav)} trial'],
        [4, 'topic_snapshot.csv', 'Data CSV', '2.7 Integrasi Sensor',
         'OTOMATIS' if snap else 'BELUM', f'{len(snap)} topik'],
        [5, 'Screenshot RViz /scan', 'Gambar', '2.7.3 Validasi Sensor', 'MANUAL', ''],
        [6, 'Screenshot RTAB-Map loop closure', 'Gambar', '2.13.3 Loc Lock', 'MANUAL', ''],
        [7, 'Screenshot RViz Nav2 path+costmap', 'Gambar', '2.14.3 Demo Nav', 'MANUAL', ''],
        [8, 'Video robot otonom', 'Video', '2.14.3 Demo Nav', 'MANUAL', ''],
        [9, 'Foto sensor terpasang', 'Gambar', '2.3.2 Penempatan', 'MANUAL', ''],
    ]
    for row in checklist:
        ws.append(row)
        st = ws.cell(row=ws.max_row, column=5)
        st.fill = GOOD if row[4] == 'OTOMATIS' else OK
    autosize(ws, maxw=34)

    # ---- INFO sheet di depan ----
    ws = wb.create_sheet('INFO', 0)
    ws['A1'] = 'AMR — Rekap Pengujian Odometry / Localization / Navigation'
    ws['A1'].font = TITLE
    ws.append([])
    ws.append([f'Generated: {now_iso()}'])
    ws.append([f'Odometry trials   : {len(odom)}'])
    ws.append([f'Localization trials: {len(loc)}'])
    ws.append([f'Navigation trials : {len(nav)}'])
    ws.append([f'Topic snapshot    : {len(snap)} topik'])
    ws.append([])
    ws.append(['Sheet _Summary = tabel siap salin ke laporan.'])
    ws.append(['Sheet _Raw     = data mentah lengkap (semua kolom).'])
    ws.append(['Evidence_Index = checklist bukti (otomatis + manual).'])
    autosize(ws, maxw=60)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--input', default='~/amr_test_results')
    args = ap.parse_args()
    indir = ensure_dir(args.input)

    odom = read_csv(os.path.join(indir, 'odometry_trials.csv'))
    loc  = read_csv(os.path.join(indir, 'localization_trials.csv'))
    nav  = read_csv(os.path.join(indir, 'navigation_trials.csv'))
    snap = read_csv(os.path.join(indir, 'topic_snapshot.csv'))

    n = write_master_csv(os.path.join(indir, 'amr_test_recap.csv'),
                         odom, loc, nav, snap)
    if n == 0:
        print('Belum ada data CSV yang bisa direkap di', indir)
        print('Jalankan logger pengujian dulu (odom/localization/navigation).')
        return

    print('=' * 60)
    print(' REKAP PENGUJIAN AMR')
    print('=' * 60)
    print(f' Input : {indir}')
    print(f' CSV   : {os.path.join(indir, "amr_test_recap.csv")}  ({n} baris)')

    xlsx = build_excel(os.path.join(indir, 'amr_test_recap.xlsx'),
                       odom, loc, nav, snap,
                       raw_dir=os.path.join(indir, 'raw'))
    if xlsx:
        print(f' XLSX  : {xlsx}')
        print('         → buka di Excel, sheet *_Summary siap salin ke laporan')
    else:
        print(' XLSX  : DILEWATI (openpyxl belum terpasang)')
        print('         pip3 install openpyxl  → jalankan ulang untuk .xlsx')

    print('-' * 60)
    print(f' Odometry trials    : {len(odom)}')
    print(f' Localization trials: {len(loc)}')
    print(f' Navigation trials  : {len(nav)}')
    print(f' Topic snapshot     : {len(snap)} topik')
    print('=' * 60)


if __name__ == '__main__':
    main()
