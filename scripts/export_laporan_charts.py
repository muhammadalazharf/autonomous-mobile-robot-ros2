#!/usr/bin/env python3
"""
export_laporan_charts.py
========================
Bangun struktur folder docs/laporan/ untuk Bab 2.10-2.12, isi dengan:
  - Copy file ANALISIS terpilih dari ~/hasil_olahan/
  - Render chart penting jadi PNG (matplotlib) supaya bisa langsung
    dipaste ke Word/Google Docs
  - README mapping ke setiap sub-bab

Output:
  docs/laporan/
  ├── README.md                                ← mapping data → sub-bab
  ├── bab_2_10_2_11_mapping/
  │   ├── PERBANDINGAN_SEMUA_ANALISIS.xlsx
  │   ├── lab_demo_18jun_ANALISIS.xlsx
  │   ├── lintasan_2d_<db>.png                ← chart 1 per peta utama
  │   ├── drift_z_<db>.png
  │   ├── yaw_vs_t_<db>.png
  │   ├── kecepatan_instan_<db>.png
  │   └── loop_closure_hist_<db>.png
  ├── bab_2_12_remapping/
  │   ├── sebelum_lab_acuan.png               ← peta awal kurang baik
  │   ├── sesudah_lab_demo_18jun.png          ← peta acuan akhir
  │   ├── bar_lc_rate_semua_db.png
  │   ├── bar_kecepatan_semua_db.png
  │   ├── bar_panjang_lintasan_semua_db.png
  │   └── bar_drift_z_semua_db.png
  └── statistik_ringkas.csv                   ← angka untuk dipaste di tabel

Default focus database (bisa diubah pakai --focus):
  - lab_demo_18jun        ← PETA ACUAN
  - lab_demo_17jun        ← perbandingan (sesi terpanjang)
  - lab_acuan             ← contoh peta awal (LC=0)
  - lab_final_20260609_212523 ← LC rate tertinggi

Dependensi:
  pip3 install openpyxl matplotlib

Cara pakai:
  python3 scripts/export_laporan_charts.py
  python3 scripts/export_laporan_charts.py \
      --input ~/hasil_olahan \
      --output docs/laporan
"""
import argparse
import csv
import math
import os
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("[ERROR] openpyxl belum terpasang. Install: pip3 install openpyxl")
    sys.exit(1)

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except ImportError:
    print("[ERROR] matplotlib belum terpasang. Install: pip3 install matplotlib")
    sys.exit(1)


FOCUS_DEFAULT = [
    'lab_demo_18jun',
    'lab_demo_17jun',
    'lab_acuan',
    'lab_final_20260609_212523',
]


def read_sheet(xlsx_path, sheet_name):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if r is None or all(v is None for v in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def render_per_db(analysis_xlsx, out_dir):
    """Render 5 chart PNG dari satu file _ANALISIS.xlsx."""
    db_name = Path(analysis_xlsx).stem.replace('_ANALISIS', '')
    traj = read_sheet(analysis_xlsx, 'TRAJECTORY_PLUS')
    hist = read_sheet(analysis_xlsx, 'LOOP_CLOSURE_HIST')

    if not traj:
        print(f"    [SKIP] {db_name}: TRAJECTORY_PLUS kosong")
        return []

    t  = [safe_float(r.get('t_rel_s')) for r in traj]
    x  = [safe_float(r.get('x_m')) for r in traj]
    y  = [safe_float(r.get('y_m')) for r in traj]
    z  = [safe_float(r.get('z_m')) for r in traj]
    yaw = [safe_float(r.get('yaw_deg')) for r in traj]
    v   = [safe_float(r.get('v_instan_mps')) for r in traj]

    files = []

    # 1. Lintasan 2D
    fig, ax = plt.subplots(figsize=(10, 8))
    ax.plot(x, y, '-', linewidth=1.2, color='#1F4E79')
    ax.plot(x[0], y[0], 'o', markersize=10, color='green', label='Start')
    ax.plot(x[-1], y[-1], '^', markersize=10, color='red', label='End')
    ax.set_xlabel('x (m)')
    ax.set_ylabel('y (m)')
    ax.set_title(f'Lintasan Mapping 2D — {db_name}')
    ax.axis('equal')
    ax.grid(True, alpha=0.3)
    ax.legend()
    p = out_dir / f'lintasan_2d_{db_name}.png'
    fig.savefig(p, dpi=150, bbox_inches='tight')
    plt.close(fig)
    files.append(p)

    # 2. Drift Z vs t
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(t, z, '-', linewidth=1.0, color='#C0504D')
    ax.set_xlabel('t (s)')
    ax.set_ylabel('z (m)')
    z_range = max(z) - min(z) if z else 0
    ax.set_title(f'Drift Z vs Waktu — {db_name}  (range = {z_range:.3f} m)')
    ax.grid(True, alpha=0.3)
    ax.axhline(0, color='gray', linewidth=0.5)
    p = out_dir / f'drift_z_{db_name}.png'
    fig.savefig(p, dpi=150, bbox_inches='tight')
    plt.close(fig)
    files.append(p)

    # 3. Yaw vs t
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(t, yaw, '-', linewidth=1.0, color='#4F81BD')
    ax.set_xlabel('t (s)')
    ax.set_ylabel('yaw (°)')
    ax.set_title(f'Orientasi Yaw vs Waktu — {db_name}')
    ax.grid(True, alpha=0.3)
    p = out_dir / f'yaw_vs_t_{db_name}.png'
    fig.savefig(p, dpi=150, bbox_inches='tight')
    plt.close(fig)
    files.append(p)

    # 4. Kecepatan instan vs t
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(t, v, '-', linewidth=0.7, color='#9BBB59', alpha=0.7)
    ax.set_xlabel('t (s)')
    ax.set_ylabel('v (m/s)')
    v_max_v = max(v) if v else 0
    ax.set_title(f'Kecepatan Instan vs Waktu — {db_name}  (v_max = {v_max_v:.3f} m/s)')
    ax.grid(True, alpha=0.3)
    p = out_dir / f'kecepatan_instan_{db_name}.png'
    fig.savefig(p, dpi=150, bbox_inches='tight')
    plt.close(fig)
    files.append(p)

    # 5. Histogram loop closure (kalau ada)
    if hist:
        top = hist[:30]
        labels = [str(r.get('from_id')) for r in top]
        counts = [int(safe_float(r.get('jumlah_LC_keluar'))) for r in top]
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.bar(range(len(top)), counts, color='#1F4E79')
        ax.set_xticks(range(len(top)))
        ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=7)
        ax.set_xlabel('from_id (node sumber)')
        ax.set_ylabel('jumlah loop closure')
        ax.set_title(f'Top-30 Loop Closure per from_id — {db_name}')
        ax.grid(True, alpha=0.3, axis='y')
        p = out_dir / f'loop_closure_hist_{db_name}.png'
        fig.savefig(p, dpi=150, bbox_inches='tight')
        plt.close(fig)
        files.append(p)

    print(f"    [✓] {db_name}: {len(files)} PNG")
    return files


def render_perbandingan(perb_xlsx, out_dir):
    """Bar chart perbandingan semua database."""
    rows = read_sheet(perb_xlsx, 'PERBANDINGAN')
    if not rows:
        print("    [SKIP] PERBANDINGAN kosong")
        return []

    rows = sorted(rows, key=lambda r: -safe_float(r.get('lc_rate_persen')))
    names = [r.get('database') for r in rows]

    def bar(values, ylabel, title, fname, color='#1F4E79', sort_desc=True):
        # Sort untuk visual yang rapi
        pairs = list(zip(names, values))
        if sort_desc:
            pairs.sort(key=lambda p: -p[1])
        nm = [p[0] for p in pairs]
        vl = [p[1] for p in pairs]

        fig, ax = plt.subplots(figsize=(12, max(5, len(nm) * 0.32)))
        ax.barh(range(len(nm)), vl, color=color)
        ax.set_yticks(range(len(nm)))
        ax.set_yticklabels(nm, fontsize=8)
        ax.invert_yaxis()
        ax.set_xlabel(ylabel)
        ax.set_title(title)
        ax.grid(True, alpha=0.3, axis='x')
        for i, val in enumerate(vl):
            ax.text(val, i, f'  {val:.2f}', va='center', fontsize=7)
        p = out_dir / fname
        fig.savefig(p, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return p

    files = []
    lc      = [safe_float(r.get('lc_rate_persen')) for r in rows]
    vrata   = [safe_float(r.get('v_rata_mps'))     for r in rows]
    panjang = [safe_float(r.get('panjang_m'))      for r in rows]
    zdrift  = [safe_float(r.get('z_drift_m'))      for r in rows]

    files.append(bar(lc, 'Loop Closure Rate (%)',
                     'Perbandingan Loop Closure Rate per Database',
                     'bar_lc_rate_semua_db.png', color='#1F4E79'))
    files.append(bar(vrata, 'Kecepatan rata-rata (m/s)',
                     'Perbandingan Kecepatan Mapping per Database',
                     'bar_kecepatan_semua_db.png', color='#9BBB59'))
    files.append(bar(panjang, 'Panjang lintasan (m)',
                     'Perbandingan Panjang Lintasan per Database',
                     'bar_panjang_lintasan_semua_db.png', color='#4F81BD'))
    files.append(bar(zdrift, 'Drift Z (m) — lebih kecil lebih baik',
                     'Perbandingan Drift Z per Database',
                     'bar_drift_z_semua_db.png', color='#C0504D'))
    return files


def write_statistik_csv(perb_xlsx, csv_path):
    rows = read_sheet(perb_xlsx, 'PERBANDINGAN')
    if not rows:
        return
    rows = sorted(rows, key=lambda r: -safe_float(r.get('lc_rate_persen')))
    headers = ['database', 'total_node', 'durasi_s', 'panjang_m',
               'v_rata_mps', 'v_max_inst', 'density_node_per_m',
               'loop_closure', 'lc_rate_persen', 'kualitas', 'z_drift_m']
    with open(csv_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, '') for h in headers])


README_MD = """# Data untuk Laporan Bab 2.10 – 2.12

Folder ini berisi data terolah dari 26 database RTAB-Map (`~/maps/*.db`)
yang sudah dianalisis menjadi metrik turunan + chart PNG siap-pakai
untuk dipaste ke Word/Google Docs.

## Sumber Data
- Ekstraksi struktur: `scripts/export_rtabmap_db.py` → `~/hasil_ekstrak/`
- Analisis turunan : `scripts/analyze_rtabmap_excel.py` → `~/hasil_olahan/`
- Render PNG      : `scripts/export_laporan_charts.py` → `docs/laporan/` (folder ini)

## Mapping Data → Sub-Bab Laporan

### Bab 2.10 — Mapping Menggunakan RTAB-Map

| Sub-bab | File yang dipakai |
|---|---|
| **2.10.1** Akuisisi Data Mapping | `bab_2_10_2_11_mapping/lab_demo_18jun_ANALISIS.xlsx` → sheet RINGKASAN_ANALISIS (durasi, panjang lintasan) |
| **2.10.2** Pembentukan Pose Graph | `bab_2_10_2_11_mapping/lintasan_2d_lab_demo_18jun.png` (Gambar utama) |
| **2.10.3** Penyimpanan Hasil Mapping | `statistik_ringkas.csv` (tabel daftar 20 database + ukuran/jumlah node) |

### Bab 2.11 — Analisis Database Mapping

| Sub-bab | File yang dipakai |
|---|---|
| **2.11.1** Identifikasi Status Database | `statistik_ringkas.csv` + catatan manual: total 26, valid 20, kosong 2, malformed 4 |
| **2.11.2** Analisis Node/Keyframe/Link/LC | `bab_2_10_2_11_mapping/loop_closure_hist_lab_demo_18jun.png` + RINGKASAN_ANALISIS |
| **2.11.3** Deduplikasi + Database Utama | `bab_2_12_remapping/bar_lc_rate_semua_db.png` (lab_demo_18jun ranking #2 dari 20) |

### Bab 2.12 — Remapping dan Pemilihan Peta Acuan

| Sub-bab | File yang dipakai |
|---|---|
| **2.12.1** Evaluasi Kualitas Peta Awal | `bab_2_12_remapping/sebelum_lab_acuan.png` (LC=0%, lintasan stuck di asal) |
| **2.12.2** Pelaksanaan Remapping | `bab_2_12_remapping/sesudah_lab_demo_18jun.png` (LC=70%, lintasan lengkap) |
| **2.12.3** Penentuan Peta Acuan | `bab_2_12_remapping/bar_lc_rate_semua_db.png` + `bar_kecepatan_semua_db.png` |

## Klasifikasi Kualitas (Loop Closure Rate)

| LC Rate | Kualitas | Interpretasi |
|---|---|---|
| > 30% | SANGAT BAIK | Banyak revisit, peta konsisten |
| 10-30% | BAIK | Cukup revisit |
| 1-10% | CUKUP | Sedikit loop closure |
| < 1% | KURANG | Mapping linear, peta rawan drift |

## Top-5 Database (dari `statistik_ringkas.csv`)

Lihat baris teratas `statistik_ringkas.csv` — sudah diurutkan berdasarkan
LC rate descending. Peta acuan **`lab_demo_18jun`** dipilih karena:
- 1846 node (terlengkap di antara peta dengan LC tinggi)
- LC rate **70.21%** (kategori SANGAT BAIK)
- Drift Z kecil, lintasan menjangkau seluruh lab

## Cara Update

Setelah ada database `.db` baru di `~/maps/`:
```bash
python3 scripts/export_rtabmap_db.py ~/maps --out ~/hasil_ekstrak
python3 scripts/analyze_rtabmap_excel.py --input ~/hasil_ekstrak --output ~/hasil_olahan
python3 scripts/export_laporan_charts.py --input ~/hasil_olahan --output docs/laporan
```
"""


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--input',  default=os.path.expanduser('~/hasil_olahan'),
                    help='Folder hasil_olahan/ (default: ~/hasil_olahan)')
    ap.add_argument('--output',
                    default=os.path.join(os.path.dirname(os.path.dirname(
                        os.path.abspath(__file__))), 'docs', 'laporan'),
                    help='Folder output (default: <repo>/docs/laporan)')
    ap.add_argument('--focus', nargs='*', default=FOCUS_DEFAULT,
                    help='Nama database fokus (tanpa _ANALISIS suffix)')
    args = ap.parse_args()

    in_dir  = Path(os.path.expanduser(args.input))
    out_dir = Path(os.path.expanduser(args.output))
    mapping_dir = out_dir / 'bab_2_10_2_11_mapping'
    remap_dir   = out_dir / 'bab_2_12_remapping'
    mapping_dir.mkdir(parents=True, exist_ok=True)
    remap_dir.mkdir(parents=True, exist_ok=True)

    perb_xlsx = in_dir / 'PERBANDINGAN_SEMUA_ANALISIS.xlsx'
    if not perb_xlsx.exists():
        print(f"[ERROR] {perb_xlsx} tidak ada — jalankan analyze_rtabmap_excel.py dulu")
        sys.exit(1)

    print('=' * 70)
    print('  EXPORT CHART LAPORAN — copy xlsx + render PNG')
    print(f'  Input : {in_dir}')
    print(f'  Output: {out_dir}')
    print(f'  Focus : {", ".join(args.focus)}')
    print('=' * 70)

    # Copy file utama
    print('\n[1/4] Copy file ANALISIS terpilih...')
    shutil.copy2(perb_xlsx, remap_dir / perb_xlsx.name)
    print(f'    [✓] {perb_xlsx.name} → {remap_dir.name}/')

    for db in args.focus:
        src = in_dir / f'{db}_ANALISIS.xlsx'
        if src.exists():
            shutil.copy2(src, mapping_dir / src.name)
            print(f'    [✓] {src.name} → {mapping_dir.name}/')
        else:
            print(f'    [!] {src.name} tidak ada — dilewati')

    # Render chart per database fokus
    print('\n[2/4] Render chart per database fokus...')
    for db in args.focus:
        src = in_dir / f'{db}_ANALISIS.xlsx'
        if not src.exists():
            continue
        # peta acuan & perbandingan → mapping folder
        # peta awal kurang baik (lab_acuan) → remap folder dengan nama "sebelum_"
        # peta acuan utama → remap folder dengan nama "sesudah_"
        render_per_db(src, mapping_dir)

    # Buat sebelum/sesudah khusus untuk Bab 2.12
    print('\n[3/4] Sebelum/sesudah Bab 2.12...')
    before_src = in_dir / 'lab_acuan_ANALISIS.xlsx'
    after_src  = in_dir / 'lab_demo_18jun_ANALISIS.xlsx'

    def render_compact(xlsx, out_png, title):
        rows = read_sheet(xlsx, 'TRAJECTORY_PLUS')
        if not rows:
            return False
        x = [safe_float(r.get('x_m')) for r in rows]
        y = [safe_float(r.get('y_m')) for r in rows]
        ring = {r.get('parameter'): r.get('nilai')
                for r in read_sheet(xlsx, 'RINGKASAN_ANALISIS')}
        lc_rate = ring.get('LOOP CLOSURE RATE', 0)
        kualitas = ring.get('kualitas_peta', '')
        panjang = ring.get('panjang_lintasan', 0)

        fig, ax = plt.subplots(figsize=(8, 8))
        ax.plot(x, y, '-', linewidth=1.5, color='#1F4E79')
        ax.plot(x[0], y[0], 'o', markersize=12, color='green', label='Start')
        ax.plot(x[-1], y[-1], '^', markersize=12, color='red', label='End')
        ax.set_xlabel('x (m)')
        ax.set_ylabel('y (m)')
        ax.set_title(f'{title}\nLC rate: {lc_rate}% — {kualitas} — lintasan {panjang} m',
                     fontsize=10)
        ax.axis('equal')
        ax.grid(True, alpha=0.3)
        ax.legend()
        fig.savefig(out_png, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return True

    if before_src.exists():
        if render_compact(before_src, remap_dir / 'sebelum_lab_acuan.png',
                          'SEBELUM Remapping — lab_acuan.db'):
            print('    [✓] sebelum_lab_acuan.png')
    if after_src.exists():
        if render_compact(after_src, remap_dir / 'sesudah_lab_demo_18jun.png',
                          'SESUDAH Remapping — lab_demo_18jun.db (PETA ACUAN)'):
            print('    [✓] sesudah_lab_demo_18jun.png')

    # Bar chart perbandingan + statistik csv
    print('\n[4/4] Bar chart perbandingan + tabel ringkas...')
    files = render_perbandingan(perb_xlsx, remap_dir)
    for f in files:
        print(f'    [✓] {f.name}')

    write_statistik_csv(perb_xlsx, out_dir / 'statistik_ringkas.csv')
    print(f'    [✓] statistik_ringkas.csv')

    # README
    (out_dir / 'README.md').write_text(README_MD)
    print(f'    [✓] README.md')

    print('\n' + '=' * 70)
    print(f'  SELESAI — output di {out_dir}')
    print(f'  Total PNG: {len(list(out_dir.glob("**/*.png")))}')
    print(f'  Total XLSX: {len(list(out_dir.glob("**/*.xlsx")))}')
    print('=' * 70)


if __name__ == '__main__':
    main()
