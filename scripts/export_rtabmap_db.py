#!/usr/bin/env python3
"""
export_rtabmap_db.py
====================
Ekstrak data dari semua file database RTAB-Map (.db) ke Excel/CSV.

TIDAK memerlukan ROS aktif — hanya Python standar + openpyxl (opsional).
File .db adalah SQLite database, bisa dibuka langsung.

Data yang diekstrak per file .db:
  Sheet RINGKASAN    : statistik database (node, durasi, loop closure count)
  Sheet TRAJECTORY   : lintasan robot (x, y, z, yaw setiap node)
  Sheet LOOP_CLOSURE : daftar semua loop closure (from→to, tipe)
  Sheet NODE_DETAIL  : detail tiap node (id, weight, timestamp, map_id)

Output ke folder hasil_ekstrak/:
  hasil_ekstrak/
  ├── RINGKASAN_SEMUA.xlsx   ← semua database dalam satu tabel ringkasan
  ├── RINGKASAN_SEMUA.csv
  ├── <db1_name>.xlsx        ← data lengkap per database (multi-sheet)
  ├── <db2_name>.xlsx
  ├── ...
  └── csv/
      ├── <db1_name>/
      │   ├── RINGKASAN.csv
      │   ├── TRAJECTORY.csv
      │   ├── LOOP_CLOSURE.csv
      │   └── NODE_DETAIL.csv
      └── ...

------------------------------------------------------------------
CARA PAKAI (di NUC, terminal biasa — TANPA perlu source ROS):

  # Proses semua .db di ~/maps/:
  python3 ~/amr_starter/scripts/export_rtabmap_db.py

  # Tentukan folder lain:
  python3 ~/amr_starter/scripts/export_rtabmap_db.py ~/maps --out ~/hasil_ekstrak

  # Proses satu file saja:
  python3 ~/amr_starter/scripts/export_rtabmap_db.py \
      --single ~/maps/lab_demo_18jun.db

Dependensi opsional untuk .xlsx:
  pip3 install openpyxl
  (tanpa openpyxl tetap menghasilkan CSV — buka langsung di Excel.)
------------------------------------------------------------------
"""

import argparse
import csv
import math
import os
import sqlite3
import struct
import sys
from pathlib import Path


# ----------------------------------------------------------------
# Decode RTAB-Map Transform blob (3×4 float32, 48 bytes, row-major)
# Layout: [r00 r01 r02 tx | r10 r11 r12 ty | r20 r21 r22 tz]
# ----------------------------------------------------------------
def decode_transform(blob):
    """Kembalikan (x, y, z, roll_deg, pitch_deg, yaw_deg) atau None."""
    if blob is None or len(blob) < 48:
        return None
    try:
        v = struct.unpack_from('12f', blob, 0)
        tx, ty, tz = v[3], v[7], v[11]
        r00, r10, r20 = v[0], v[4], v[8]
        r21, r22 = v[9], v[10]
        yaw   = math.atan2(r10, r00)
        pitch = math.atan2(-r20, math.sqrt(r21 ** 2 + r22 ** 2))
        roll  = math.atan2(r21, r22)
        return (round(tx, 4), round(ty, 4), round(tz, 4),
                round(math.degrees(roll), 2),
                round(math.degrees(pitch), 2),
                round(math.degrees(yaw), 2))
    except Exception:
        return None


LINK_TYPE = {
    0: 'Neighbor',
    1: 'NeighborMerged',
    2: 'GlobalLoopClosure',
    3: 'LocalLoopClosure',
    4: 'User',
    5: 'Virtual',
    6: 'LocalSpaceClosure',
    7: 'Landmark',
    8: 'Gravity',
}


def table_names(conn):
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    return {r[0] for r in cur.fetchall()}


def column_names(conn, table):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {r[1] for r in cur.fetchall()}


# ----------------------------------------------------------------
# Ekstrak satu .db
# ----------------------------------------------------------------
def extract_db(db_path, out_dir):
    db_name = Path(db_path).stem
    print(f"\n[→] {Path(db_path).name}  ({os.path.getsize(db_path)/1e6:.1f} MB)")

    try:
        conn = sqlite3.connect(f'file:{db_path}?mode=ro', uri=True)
    except Exception as e:
        print(f"    [ERROR] Tidak bisa buka: {e}")
        return None

    tables = table_names(conn)
    cur = conn.cursor()

    summary = {
        'database':            db_name,
        'ukuran_MB':           round(os.path.getsize(db_path) / 1e6, 1),
        'total_node':          0,
        'jumlah_map':          0,
        'durasi_s':            0.0,
        'lintasan_panjang_m':  0.0,
        'loop_closure_global': 0,
        'loop_closure_lokal':  0,
        'neighbor_links':      0,
    }

    # ---- Node / Trajectory ----
    trajectory_rows = []
    node_detail_rows = []
    t0 = None

    if 'Node' in tables:
        cols = column_names(conn, 'Node')
        select_cols = 'id, map_id, weight, stamp, pose'
        select_cols += ', label' if 'label' in cols else ", '' as label"
        try:
            cur.execute(f"SELECT {select_cols} FROM Node ORDER BY stamp")
            nodes = cur.fetchall()
            summary['total_node'] = len(nodes)

            if nodes:
                summary['jumlah_map'] = len({n[1] for n in nodes})
                stamps = [n[3] for n in nodes if n[3] and n[3] > 0]
                if stamps:
                    t0 = min(stamps)
                    summary['durasi_s'] = round(max(stamps) - t0, 2)

            prev_xy = None
            total_dist = 0.0

            for row in nodes:
                nid, map_id, weight, stamp, pose_blob, label = row
                rel_t = round(stamp - t0, 4) if t0 and stamp else 0.0
                dec = decode_transform(pose_blob)

                if dec:
                    x, y, z, roll, pitch, yaw = dec
                    trajectory_rows.append(
                        [nid, rel_t, x, y, z, roll, pitch, yaw])
                    if prev_xy:
                        total_dist += math.hypot(x - prev_xy[0], y - prev_xy[1])
                    prev_xy = (x, y)

                node_detail_rows.append(
                    [nid, map_id, weight,
                     round(stamp, 4) if stamp else '',
                     rel_t, label or ''])

            summary['lintasan_panjang_m'] = round(total_dist, 2)
        except Exception as e:
            print(f"    [WARN] Error baca Node: {e}")

    # ---- Link / Loop Closure ----
    link_rows = []
    if 'Link' in tables:
        try:
            cur.execute("SELECT from_id, to_id, type FROM Link ORDER BY from_id")
            for from_id, to_id, ltype in cur.fetchall():
                tname = LINK_TYPE.get(ltype, f'Type{ltype}')
                link_rows.append([from_id, to_id, ltype, tname])
                if ltype == 2:
                    summary['loop_closure_global'] += 1
                elif ltype == 3:
                    summary['loop_closure_lokal'] += 1
                elif ltype == 0:
                    summary['neighbor_links'] += 1
        except Exception as e:
            print(f"    [WARN] Error baca Link: {e}")

    conn.close()

    # ---- Susun sheet data ----
    sheets = {
        'RINGKASAN': {
            'headers': ['parameter', 'nilai'],
            'rows': [
                ['database',            summary['database']],
                ['ukuran_file_MB',      summary['ukuran_MB']],
                ['total_node',          summary['total_node']],
                ['jumlah_map_id',       summary['jumlah_map']],
                ['durasi_recording_s',  summary['durasi_s']],
                ['panjang_lintasan_m',  summary['lintasan_panjang_m']],
                ['loop_closure_global', summary['loop_closure_global']],
                ['loop_closure_lokal',  summary['loop_closure_lokal']],
                ['neighbor_links',      summary['neighbor_links']],
            ],
        },
        'TRAJECTORY': {
            'headers': ['node_id', 't_rel_s', 'x_m', 'y_m', 'z_m',
                        'roll_deg', 'pitch_deg', 'yaw_deg'],
            'rows': trajectory_rows,
        },
        'LOOP_CLOSURE': {
            'headers': ['from_id', 'to_id', 'type_id', 'type_nama'],
            'rows': link_rows,
        },
        'NODE_DETAIL': {
            'headers': ['node_id', 'map_id', 'weight', 'stamp_s', 't_rel_s', 'label'],
            'rows': node_detail_rows,
        },
    }

    # ---- Tulis CSV (selalu) ----
    csv_subdir = os.path.join(out_dir, 'csv', db_name)
    os.makedirs(csv_subdir, exist_ok=True)
    for sname, sd in sheets.items():
        with open(os.path.join(csv_subdir, f'{sname}.csv'), 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(sd['headers'])
            w.writerows(sd['rows'])

    # ---- Tulis Excel (jika openpyxl ada) ----
    xlsx_path = None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
        HEADER_FONT  = Font(bold=True, color='FFFFFF')
        HEADER_ALIGN = Alignment(horizontal='center')

        wb = Workbook()
        first = True
        for sname, sd in sheets.items():
            ws = wb.active if first else wb.create_sheet(title=sname)
            if first:
                ws.title = sname
                first = False
            ws.append(sd['headers'])
            for cell in ws[1]:
                cell.font  = HEADER_FONT
                cell.fill  = HEADER_FILL
                cell.alignment = HEADER_ALIGN
            for row in sd['rows']:
                ws.append(row)
            for col in ws.columns:
                w = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 35)

        xlsx_path = os.path.join(out_dir, f'{db_name}.xlsx')
        wb.save(xlsx_path)
        print(f"    [✓] {db_name}.xlsx  —  "
              f"{summary['total_node']} node | "
              f"{summary['lintasan_panjang_m']} m | "
              f"LC {summary['loop_closure_global']} global / "
              f"{summary['loop_closure_lokal']} lokal | "
              f"{summary['durasi_s']} s")
    except ImportError:
        print(f"    [✓] CSV → {csv_subdir}/  "
              f"(openpyxl tidak ada; pip3 install openpyxl untuk .xlsx)")

    return summary, xlsx_path


# ----------------------------------------------------------------
# main
# ----------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('maps_dir', nargs='?',
                    default=os.path.expanduser('~/maps'),
                    help='Folder berisi file .db (default: ~/maps)')
    ap.add_argument('--out', default=os.path.expanduser('~/hasil_ekstrak'),
                    help='Folder output (default: ~/hasil_ekstrak)')
    ap.add_argument('--single',
                    help='Proses satu file .db saja, abaikan maps_dir')
    args = ap.parse_args()

    out_dir = os.path.expanduser(args.out)
    os.makedirs(out_dir, exist_ok=True)

    if args.single:
        db_files = [Path(os.path.expanduser(args.single))]
    else:
        maps_dir = os.path.expanduser(args.maps_dir)
        if not os.path.isdir(maps_dir):
            print(f"[ERROR] Folder tidak ditemukan: {maps_dir}")
            sys.exit(1)
        db_files = sorted(Path(maps_dir).glob('*.db'))
        if not db_files:
            print(f"[ERROR] Tidak ada file .db di: {maps_dir}")
            sys.exit(1)

    print('=' * 65)
    print(f'  EXPORT RTAB-MAP DB → EXCEL/CSV')
    print(f'  Input : {args.single or args.maps_dir}')
    print(f'  Output: {out_dir}')
    print(f'  Total : {len(db_files)} file .db ditemukan')
    print('=' * 65)

    all_summaries = []
    for db_file in db_files:
        result = extract_db(str(db_file), out_dir)
        if result:
            summary, _ = result
            all_summaries.append(summary)

    if not all_summaries:
        print("\n[ERROR] Tidak ada data berhasil diekstrak.")
        sys.exit(1)

    # ---- Ringkasan semua db ----
    ring_headers = ['database', 'ukuran_MB', 'total_node', 'jumlah_map',
                    'durasi_s', 'panjang_lintasan_m',
                    'loop_closure_global', 'loop_closure_lokal', 'neighbor_links']
    ring_rows = [[s['database'], s['ukuran_MB'], s['total_node'],
                  s['jumlah_map'], s['durasi_s'], s['lintasan_panjang_m'],
                  s['loop_closure_global'], s['loop_closure_lokal'],
                  s['neighbor_links']] for s in all_summaries]

    # CSV ringkasan (selalu)
    ring_csv = os.path.join(out_dir, 'RINGKASAN_SEMUA.csv')
    with open(ring_csv, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(ring_headers)
        w.writerows(ring_rows)

    # Excel ringkasan
    ring_path = ring_csv
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'RINGKASAN_SEMUA'
        ws.append(ring_headers)
        for cell in ws[1]:
            cell.font  = Font(bold=True, color='FFFFFF')
            cell.fill  = PatternFill('solid', fgColor='1F4E79')
            cell.alignment = Alignment(horizontal='center')
        for row in ring_rows:
            ws.append(row)
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 35)

        ring_path = os.path.join(out_dir, 'RINGKASAN_SEMUA.xlsx')
        wb.save(ring_path)
    except ImportError:
        pass

    # ---- Laporan akhir ----
    print('\n' + '=' * 65)
    print(f'  SELESAI — {len(all_summaries)} database berhasil diekstrak')
    print(f'  Ringkasan: {ring_path}')
    print(f'\n  {"Database":<32} {"Node":>6} {"Panjang":>9} {"LC":>4} {"Durasi":>9}')
    print(f'  {"-"*32} {"-"*6} {"-"*9} {"-"*4} {"-"*9}')
    for s in all_summaries:
        print(f'  {s["database"][:32]:<32} {s["total_node"]:>6} '
              f'{s["lintasan_panjang_m"]:>8.1f}m '
              f'{s["loop_closure_global"]:>4} '
              f'{s["durasi_s"]:>8.1f}s')
    print('=' * 65)
    print(f'\n  Folder output: {out_dir}')


if __name__ == '__main__':
    main()
